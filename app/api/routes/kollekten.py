"""GET /api/kollekten  —  Tabellendaten aus kollekten_uebersicht.xlsx."""
from __future__ import annotations

import sys
from pathlib import Path
from typing import Optional

from fastapi import APIRouter, Query

router = APIRouter()

ROOT = Path(__file__).parent.parent.parent.parent


def _load_cfg() -> dict:
    sys.path.insert(0, str(ROOT))
    from config import load_config
    return load_config()


def _eur(v: float) -> str:
    return "{:,.2f}".format(v).replace(",", "X").replace(".", ",").replace("X", ".")


def _load_rows(cfg: dict) -> list[dict]:
    try:
        import openpyxl
        path = Path(cfg["output"]["overview_file"])
        if not path.exists():
            return []
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if "Uebersicht" not in wb.sheetnames:
            wb.close()
            return []
        ws = wb["Uebersicht"]
        headers = [str(ws.cell(1, c).value or "") for c in range(1, ws.max_column + 1)]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            rows.append(dict(zip(headers, row)))
        wb.close()
        return rows
    except Exception:
        return []


@router.get("/kollekten")
def get_kollekten(
    month: Optional[int] = Query(None, ge=1, le=12),
    year: Optional[int] = Query(None),
    only_warnings: bool = Query(False),
) -> dict:
    try:
        cfg = _load_cfg()
        rows = _load_rows(cfg)
        entries = []
        for record in reversed(rows):
            bd = record.get("booking_date")
            if not hasattr(bd, "month"):
                continue
            if month and bd.month != month:
                continue
            if year and bd.year != year:
                continue
            needs_review = bool(record.get("needs_review", False))
            if only_warnings and not needs_review:
                continue
            try:
                betrag = float(record.get("amount") or 0)
            except Exception:
                betrag = 0.0
            scope = str(record.get("scope") or "")
            entries.append({
                "datum": bd.strftime("%d.%m.%Y"),
                "datum_iso": bd.strftime("%Y-%m-%d"),
                "betrag": betrag,
                "betrag_fmt": _eur(betrag),
                "zweck": str(record.get("purpose") or ""),
                "typ": scope,
                "typ_label": "→ Weiterleit." if "weiter" in scope else "✓ Eigene",
                "aobj": str(record.get("aobj") or ""),
                "needs_review": needs_review,
                "datei": str(record.get("target_file") or ""),
            })
        return {"entries": entries, "count": len(entries)}
    except Exception as exc:
        return {"error": str(exc), "entries": [], "count": 0}


@router.get("/kollekten/summary")
def get_summary(
    month: Optional[int] = Query(None, ge=1, le=12),
    year: Optional[int] = Query(None),
) -> dict:
    try:
        cfg = _load_cfg()
        rows = _load_rows(cfg)
        eigene = 0.0
        weiter = 0.0
        count = 0
        for record in rows:
            bd = record.get("booking_date")
            if not hasattr(bd, "month"):
                continue
            if month and bd.month != month:
                continue
            if year and bd.year != year:
                continue
            try:
                betrag = float(record.get("amount") or 0)
            except Exception:
                betrag = 0.0
            scope = str(record.get("scope") or "")
            if "weiter" in scope:
                weiter += betrag
            else:
                eigene += betrag
            count += 1
        return {
            "summe_eigene": eigene,
            "summe_eigene_fmt": _eur(eigene),
            "summe_weiterleitung": weiter,
            "summe_weiterleitung_fmt": _eur(weiter),
            "summe_gesamt": eigene + weiter,
            "summe_gesamt_fmt": _eur(eigene + weiter),
            "count": count,
        }
    except Exception as exc:
        return {"error": str(exc)}
