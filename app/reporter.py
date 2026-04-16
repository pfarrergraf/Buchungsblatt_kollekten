"""Monatsbericht: Drucken, PDF-Export, E-Mail-Versand (Phase 4)."""
from __future__ import annotations

import sys
from datetime import date
from pathlib import Path
from typing import NamedTuple, Optional

from PySide6.QtCore import Qt
from PySide6.QtGui import QFont, QPainter, QPageSize
from PySide6.QtPrintSupport import QPrinter, QPrintDialog, QPrintPreviewDialog
from PySide6.QtWidgets import QWidget


# ── Datenmodell ───────────────────────────────────────────────────────────────

class ReportEntry(NamedTuple):
    datum: str        # "15.04.2026"
    betrag: float
    zweck: str
    typ: str          # "eigene_gemeinde" | "zur_weiterleitung"
    aobj: str


class ReportData(NamedTuple):
    monat: int
    jahr: int
    gemeinde: str
    eintraege: list[ReportEntry]
    summe_eigene: float
    summe_weiterleitung: float
    summe_gesamt: float


# ── Daten laden ───────────────────────────────────────────────────────────────

def generate_monthly_report(month: int, year: int, cfg: dict) -> ReportData:
    """Liest Kollekten-Übersicht und aggregiert einen Monatsbericht."""
    try:
        import openpyxl
        path = Path(cfg["output"]["overview_file"])
        if not path.exists():
            return _empty_report(month, year, cfg)
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if "Uebersicht" not in wb.sheetnames:
            wb.close()
            return _empty_report(month, year, cfg)
        ws = wb["Uebersicht"]
        headers = [str(ws.cell(1, c).value or "") for c in range(1, ws.max_column + 1)]
        eintraege: list[ReportEntry] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            record = dict(zip(headers, row))
            bd = record.get("booking_date")
            if not hasattr(bd, "month"):
                continue
            if bd.month != month or bd.year != year:
                continue
            try:
                betrag = float(record.get("amount") or 0)
            except Exception:
                betrag = 0.0
            eintraege.append(ReportEntry(
                datum=bd.strftime("%d.%m.%Y"),
                betrag=betrag,
                zweck=str(record.get("purpose") or ""),
                typ=str(record.get("scope") or ""),
                aobj=str(record.get("aobj") or ""),
            ))
        wb.close()
    except Exception:
        return _empty_report(month, year, cfg)

    eintraege.sort(key=lambda e: e.datum)
    summe_eigene = sum(e.betrag for e in eintraege if "weiter" not in e.typ)
    summe_weiter = sum(e.betrag for e in eintraege if "weiter" in e.typ)
    org = cfg.get("organization", {})
    gemeinde = org.get("gemeinde_name") or "RV {}".format(org.get("rechtsträger_nr", ""))
    return ReportData(
        monat=month, jahr=year, gemeinde=gemeinde,
        eintraege=eintraege,
        summe_eigene=summe_eigene,
        summe_weiterleitung=summe_weiter,
        summe_gesamt=summe_eigene + summe_weiter,
    )


def _empty_report(month: int, year: int, cfg: dict) -> ReportData:
    org = cfg.get("organization", {})
    gemeinde = org.get("gemeinde_name") or "RV {}".format(org.get("rechtsträger_nr", ""))
    return ReportData(month, year, gemeinde, [], 0.0, 0.0, 0.0)


# ── Zeichnen ──────────────────────────────────────────────────────────────────

def _eur(v: float) -> str:
    return "{:,.2f} €".format(v).replace(",", "X").replace(".", ",").replace("X", ".")


def _draw_report(painter: QPainter, printer: QPrinter, report: ReportData) -> None:
    """Zeichnet den Bericht auf den QPainter (Drucker oder PDF)."""
    rect = printer.pageRect(QPrinter.Unit.DevicePixel)
    w = int(rect.width())
    margin = int(w * 0.08)
    content_w = w - 2 * margin
    y = margin

    def text(x: int, yy: int, width: int, txt: str,
             bold: bool = False, size: int = 10,
             align: Qt.AlignmentFlag = Qt.AlignmentFlag.AlignLeft) -> int:
        font = QFont("Segoe UI", size, QFont.Weight.Bold if bold else QFont.Weight.Normal)
        painter.setFont(font)
        fm = painter.fontMetrics()
        line_h = fm.height() + 4
        painter.drawText(x, yy, width, line_h, int(align), txt)
        return line_h

    monat_name = [
        "", "Januar", "Februar", "März", "April", "Mai", "Juni",
        "Juli", "August", "September", "Oktober", "November", "Dezember",
    ][report.monat]

    # ── Kopfzeile ──
    y += text(margin, y, content_w, report.gemeinde, bold=True, size=14)
    y += 4
    y += text(margin, y, content_w, f"Kollekten-Monatsbericht {monat_name} {report.jahr}",
              bold=False, size=11)
    y += 16

    # ── Trennlinie ──
    painter.drawLine(margin, y, margin + content_w, y)
    y += 8

    # ── Spaltenbreiten ──
    col_datum = int(content_w * 0.14)
    col_betrag = int(content_w * 0.14)
    col_typ = int(content_w * 0.20)
    col_aobj = int(content_w * 0.10)
    col_zweck = content_w - col_datum - col_betrag - col_typ - col_aobj

    # ── Tabellenkopf ──
    x0 = margin
    lh = text(x0, y, col_datum, "Datum", bold=True, size=9)
    text(x0 + col_datum, y, col_betrag, "Betrag", bold=True, size=9,
         align=Qt.AlignmentFlag.AlignRight)
    text(x0 + col_datum + col_betrag + 4, y, col_zweck, "Verwendungszweck", bold=True, size=9)
    text(x0 + col_datum + col_betrag + col_zweck + 4, y, col_typ, "Typ", bold=True, size=9)
    text(x0 + content_w - col_aobj, y, col_aobj, "AObj", bold=True, size=9,
         align=Qt.AlignmentFlag.AlignRight)
    y += lh + 2
    painter.drawLine(margin, y, margin + content_w, y)
    y += 6

    # ── Tabellenzeilen ──
    page_h = int(rect.height())
    for entry in report.eintraege:
        if y + 20 > page_h - margin:
            printer.newPage()
            y = margin
        lh = text(x0, y, col_datum, entry.datum, size=9)
        text(x0 + col_datum, y, col_betrag, _eur(entry.betrag), size=9,
             align=Qt.AlignmentFlag.AlignRight)
        zweck_short = entry.zweck[:60] + ("…" if len(entry.zweck) > 60 else "")
        text(x0 + col_datum + col_betrag + 4, y, col_zweck, zweck_short, size=9)
        typ_str = "→ Weiterleit." if "weiter" in entry.typ else "✓ Eigene"
        text(x0 + col_datum + col_betrag + col_zweck + 4, y, col_typ, typ_str, size=9)
        text(x0 + content_w - col_aobj, y, col_aobj, entry.aobj, size=9,
             align=Qt.AlignmentFlag.AlignRight)
        y += lh + 2

    # ── Summenzeilen ──
    y += 8
    painter.drawLine(margin, y, margin + content_w, y)
    y += 6

    def sum_row(label: str, val: float) -> None:
        nonlocal y
        lh = text(x0, y, content_w - col_betrag, label, size=9)
        text(x0 + content_w - col_betrag, y, col_betrag, _eur(val), bold=True, size=9,
             align=Qt.AlignmentFlag.AlignRight)
        y += lh + 2

    sum_row("Summe eigene Gemeinde:", report.summe_eigene)
    sum_row("Summe Weiterleitung:", report.summe_weiterleitung)
    y += 2
    painter.drawLine(margin + content_w - col_betrag - 20, y, margin + content_w, y)
    y += 4
    sum_row("Gesamtsumme:", report.summe_gesamt)


# ── Öffentliche Funktionen ────────────────────────────────────────────────────

def print_report(report: ReportData, parent: Optional[QWidget] = None) -> bool:
    """Öffnet den Druckdialog und druckt den Bericht."""
    printer = QPrinter(QPrinter.PrinterMode.HighResolution)
    printer.setPageSize(QPageSize(QPageSize.PageSizeId.A4))
    dlg = QPrintDialog(printer, parent)
    if dlg.exec() != QPrintDialog.DialogCode.Accepted:
        return False
    painter = QPainter(printer)
    _draw_report(painter, printer, report)
    painter.end()
    return True


def preview_report(report: ReportData, parent: Optional[QWidget] = None) -> None:
    """Öffnet den Druckvorschau-Dialog."""
    printer = QPrinter(QPrinter.PrinterMode.HighResolution)
    printer.setPageSize(QPageSize(QPageSize.PageSizeId.A4))

    dlg = QPrintPreviewDialog(printer, parent)
    dlg.paintRequested.connect(
        lambda p: (_draw_report(QPainter(p), p, report), QPainter(p).end())
    )
    # Kompaktere Version:
    def _paint(p: QPrinter) -> None:
        pa = QPainter(p)
        _draw_report(pa, p, report)
        pa.end()
    dlg.paintRequested.connect(_paint)
    dlg.exec()


def export_pdf(report: ReportData, path: str) -> bool:
    """Exportiert den Bericht als PDF-Datei."""
    printer = QPrinter(QPrinter.PrinterMode.HighResolution)
    printer.setPageSize(QPageSize(QPageSize.PageSizeId.A4))
    printer.setOutputFormat(QPrinter.OutputFormat.PdfFormat)
    printer.setOutputFileName(path)
    painter = QPainter(printer)
    _draw_report(painter, printer, report)
    painter.end()
    return Path(path).exists()


def email_report(report: ReportData, cfg: dict) -> bool:
    """Versendet den Bericht per Outlook als PDF-Anhang."""
    try:
        import tempfile
        import os
        pdf_path = os.path.join(tempfile.gettempdir(),
                                "Kollekten_{:02d}_{}.pdf".format(report.monat, report.jahr))
        if not export_pdf(report, pdf_path):
            return False
        sys.path.insert(0, str(Path(__file__).parent.parent))
        from email_sender import send_results
        send_results(cfg, [Path(pdf_path)])
        return True
    except Exception:
        return False
