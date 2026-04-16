"""CollectionTable — wiederverwendbare Tabelle für verarbeitete Kollekten."""
from __future__ import annotations

import subprocess
import sys
from pathlib import Path
from typing import Optional

from PySide6.QtCore import Qt, Signal
from PySide6.QtGui import QColor, QBrush
from PySide6.QtWidgets import (
    QComboBox,
    QDialog,
    QFormLayout,
    QHeaderView,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QMenu,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
)

from booking_store import get_booking_rows
from file_actions import open_file, open_folder, reveal_in_explorer


class KorrekturDialog(QDialog):
    """Erlaubt das manuelle Überschreiben einer Klassifizierung."""

    def __init__(self, verwendungszweck: str, current_scope: str, current_aobj: str, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Klassifizierung korrigieren")
        self.setFixedSize(420, 260)
        self._verwendungszweck = verwendungszweck
        self._build_ui(current_scope, current_aobj)

    def _build_ui(self, current_scope: str, current_aobj: str):
        layout = QVBoxLayout(self)
        layout.setSpacing(12)

        lbl = QLabel(f"Verwendungszweck: <b>{self._verwendungszweck[:60]}</b>")
        lbl.setWordWrap(True)
        layout.addWidget(lbl)

        form = QFormLayout()
        form.setSpacing(8)

        self._scope_box = QComboBox()
        self._scope_box.addItem("✓ Eigene Gemeinde", "eigene_gemeinde")
        self._scope_box.addItem("→ Zur Weiterleitung", "zur_weiterleitung")
        self._scope_box.setCurrentIndex(1 if "weiter" in current_scope else 0)
        form.addRow("Buchungsart:", self._scope_box)

        self._aobj_box = QComboBox()
        self._load_aobj_codes()
        if current_aobj:
            for i in range(self._aobj_box.count()):
                if self._aobj_box.itemData(i) == current_aobj:
                    self._aobj_box.setCurrentIndex(i)
                    break
        form.addRow("AObj-Code:", self._aobj_box)

        self._reason_edit = QLineEdit()
        self._reason_edit.setPlaceholderText("Kurze Begründung (optional)")
        form.addRow("Grund:", self._reason_edit)

        layout.addLayout(form)

        hint = QLabel(
            "Diese Korrektur wird in data/reference/manual_overrides.json gespeichert\n"
            "und bei allen zukünftigen E-Mails mit diesem Verwendungszweck angewendet."
        )
        hint.setObjectName("hint")
        hint.setWordWrap(True)
        layout.addWidget(hint)

        btn_row = QHBoxLayout()
        btn_ok = QPushButton("Speichern")
        btn_ok.setObjectName("primaryButton")
        btn_ok.clicked.connect(self.accept)
        btn_cancel = QPushButton("Abbrechen")
        btn_cancel.clicked.connect(self.reject)
        btn_row.addStretch()
        btn_row.addWidget(btn_cancel)
        btn_row.addWidget(btn_ok)
        layout.addLayout(btn_row)

    def _load_aobj_codes(self):
        try:
            import json
            sys.path.insert(0, str(Path(__file__).parent.parent.parent))
            from config import get_config

            cfg = get_config()
            aobj_file = cfg.get("reference_sources", {}).get("aobj_file", "")
            if aobj_file and Path(aobj_file).exists():
                with open(aobj_file, encoding="utf-8") as f:
                    codes = json.load(f)
                for entry in codes:
                    self._aobj_box.addItem(
                        f"{entry['code']} – {entry.get('label', entry.get('bezeichnung', ''))}",
                        entry["code"],
                    )
                return
        except Exception:
            pass
        for code, label in [
            ("0110", "Gottesdienst"),
            ("0210", "Kirchenmusik"),
            ("0430", "Kinder- und Jugendarbeit"),
            ("0510", "Gemeindearbeit"),
            ("3611", "Kollekten Weiterleitung"),
        ]:
            self._aobj_box.addItem(f"{code} – {label}", code)

    def get_result(self) -> dict:
        return {
            "pattern": self._verwendungszweck,
            "scope": self._scope_box.currentData(),
            "template_kind": self._scope_box.currentData(),
            "aobj": self._aobj_box.currentData() or "",
            "reason": self._reason_edit.text().strip(),
        }


COLUMNS = [
    ("Datum", 90),
    ("Betrag", 90),
    ("Verwendungszweck", -1),
    ("Typ", 120),
    ("AObj", 60),
    ("⚠", 30),
    ("Datei", 180),
]


class CollectionTable(QTableWidget):
    """Tabelle aller verarbeiteten Kollekten mit Dateiaktionen und Mehrfachauswahl."""

    correction_saved = Signal(str, str, str)

    def __init__(self, parent=None):
        super().__init__(0, len(COLUMNS), parent)
        self._all_rows: list[dict] = []
        self._cfg = None
        self._setup_table()

    def _setup_table(self):
        self.setHorizontalHeaderLabels([c[0] for c in COLUMNS])
        for i, (_, width) in enumerate(COLUMNS):
            if width == -1:
                self.horizontalHeader().setSectionResizeMode(i, QHeaderView.ResizeMode.Stretch)
            else:
                self.horizontalHeader().setSectionResizeMode(i, QHeaderView.ResizeMode.Fixed)
                self.setColumnWidth(i, width)
        self.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.setSelectionMode(QTableWidget.SelectionMode.ExtendedSelection)
        self.setAlternatingRowColors(True)
        self.verticalHeader().setVisible(False)
        self.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.customContextMenuRequested.connect(self._show_context_menu)
        self.cellDoubleClicked.connect(self._on_double_click)

    def load_data(
        self,
        cfg: dict,
        month_filter: Optional[int] = None,
        year_filter: Optional[int] = None,
        only_warnings: bool = False,
        months_filter: Optional[set[int]] = None,
        years_filter: Optional[set[int]] = None,
    ):
        self._cfg = cfg
        self._all_rows = get_booking_rows(cfg, include_statuses={"ok"})
        if not self._all_rows:
            self._all_rows = self._load_rows_from_overview(cfg)
        self._apply_filter(
            month_filter,
            year_filter,
            only_warnings,
            months_filter=months_filter,
            years_filter=years_filter,
        )

    def _load_rows_from_overview(self, cfg: dict) -> list[dict]:
        try:
            import openpyxl

            path = Path(cfg["output"]["overview_file"])
            if not path.exists():
                return []
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            if "Uebersicht" not in wb.sheetnames:
                return []
            ws = wb["Uebersicht"]
            headers = [str(ws.cell(1, c).value or "") for c in range(1, ws.max_column + 1)]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                rows.append(dict(zip(headers, row)))
            wb.close()
            return rows
        except Exception:
            return []

    def _apply_filter(
        self,
        month: Optional[int],
        year: Optional[int],
        only_warnings: bool,
        *,
        months_filter: Optional[set[int]] = None,
        years_filter: Optional[set[int]] = None,
    ):
        self.setRowCount(0)
        warn_col = QColor("#FFF9C4")
        active_months = months_filter if months_filter else ({month} if month else set())
        active_years = years_filter if years_filter else ({year} if year else set())

        for record in reversed(self._all_rows):
            booking_date = self._coerce_date(record.get("booking_date"))
            if booking_date is None:
                continue
            if active_months and booking_date.month not in active_months:
                continue
            if active_years and booking_date.year not in active_years:
                continue
            needs_review = bool(record.get("needs_review", False))
            if only_warnings and not needs_review:
                continue

            row = self.rowCount()
            self.insertRow(row)

            self._set(row, 0, booking_date.strftime("%d.%m.%Y"))
            try:
                betrag_str = f"{float(record.get('amount') or 0):,.2f} €".replace(",", "X").replace(".", ",").replace("X", ".")
            except Exception:
                betrag_str = str(record.get("amount") or "")
            self._set(row, 1, betrag_str, align=Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self._set(row, 2, str(record.get("purpose", "") or ""))

            scope = str(record.get("scope", "") or "")
            typ_str = "→ Weiterleit." if "weiter" in scope else "✓ Eigene"
            color = QColor("#E3F2FD") if "weiter" in scope else QColor("#E8F5E9")
            item = QTableWidgetItem(typ_str)
            item.setBackground(QBrush(color))
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.setItem(row, 3, item)

            self._set(row, 4, str(record.get("aobj", "") or ""), align=Qt.AlignmentFlag.AlignCenter)

            warn_item = QTableWidgetItem("⚠" if needs_review else "")
            warn_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.setItem(row, 5, warn_item)

            target_name = Path(str(record.get("target_file", "") or "")).name
            self._set(row, 6, target_name or "—")

            if needs_review:
                for col in range(self.columnCount()):
                    item = self.item(row, col)
                    if item:
                        item.setBackground(QBrush(warn_col))

            self.item(row, 0).setData(Qt.ItemDataRole.UserRole, dict(record))

    def _coerce_date(self, value):
        if hasattr(value, "year") and hasattr(value, "month"):
            return value
        if isinstance(value, str) and value:
            try:
                from datetime import date

                return date.fromisoformat(value[:10])
            except Exception:
                return None
        return None

    def _set(self, row: int, col: int, text: str, align=None):
        item = QTableWidgetItem(text)
        if align:
            item.setTextAlignment(align)
        self.setItem(row, col, item)

    def _show_context_menu(self, pos):
        row = self.rowAt(pos.y())
        if row < 0:
            return
        item = self.item(row, 0)
        if not item:
            return
        record = item.data(Qt.ItemDataRole.UserRole)
        if not record:
            return

        menu = QMenu(self)
        act_correct = menu.addAction("Klassifizierung korrigieren…")
        act_open = menu.addAction("Datei öffnen")
        act_folder = menu.addAction("Ordner öffnen")
        act_explorer = menu.addAction("In Explorer öffnen")
        action = menu.exec(self.viewport().mapToGlobal(pos))

        if action == act_correct:
            self._open_correction(record)
        elif action == act_open:
            self.open_record_file(record)
        elif action == act_folder:
            self.open_record_folder(record)
        elif action == act_explorer:
            self._open_in_explorer(record)

    def _on_double_click(self, row: int, col: int) -> None:
        if col != 6:
            return
        item = self.item(row, 0)
        if not item:
            return
        record = item.data(Qt.ItemDataRole.UserRole)
        if record:
            self.open_record_file(record)

    def _open_correction(self, record: dict):
        dlg = KorrekturDialog(
            verwendungszweck=str(record.get("purpose", "")),
            current_scope=str(record.get("scope", "")),
            current_aobj=str(record.get("aobj", "")),
            parent=self,
        )
        if dlg.exec() == QDialog.DialogCode.Accepted:
            result = dlg.get_result()
            self._save_override(result)
            self.correction_saved.emit(result["pattern"], result["scope"], result["aobj"])

    def _save_override(self, result: dict):
        from PySide6.QtWidgets import QMessageBox
        try:
            import json

            if not self._cfg:
                return
            path = Path(
                self._cfg.get("reference_sources", {}).get(
                    "manual_overrides_file",
                    Path(__file__).parent.parent.parent / "data" / "reference" / "manual_overrides.json",
                )
            )
            existing = []
            if path.exists():
                with path.open(encoding="utf-8") as f:
                    raw = json.load(f)
                existing = raw if isinstance(raw, list) else raw.get("items", [])
            pattern = result["pattern"]
            existing = [e for e in existing if e.get("pattern") != pattern]
            existing.append(
                {
                    "pattern": pattern,
                    "scope": result["scope"],
                    "template_kind": result.get("template_kind", result["scope"]),
                    "aobj": result["aobj"],
                    "reason": result.get("reason", "Manuelle Korrektur"),
                    "confidence": 1.0,
                }
            )
            path.parent.mkdir(parents=True, exist_ok=True)
            with path.open("w", encoding="utf-8") as f:
                json.dump(existing, f, indent=2, ensure_ascii=False)
        except Exception as exc:
            QMessageBox.warning(
                self,
                "Korrektur konnte nicht gespeichert werden",
                f"Die Klassifizierungskorrektur wurde NICHT gespeichert.\n\nFehler: {exc}",
            )

    def _open_in_explorer(self, record: dict):
        target = str(record.get("target_file", "") or "")
        if target and reveal_in_explorer(target):
            return
        if self._cfg:
            out_dir = self._cfg.get("output", {}).get("root_dir", "")
            if out_dir:
                subprocess.Popen(["explorer", out_dir])

    def open_record_file(self, record: dict) -> bool:
        target = str(record.get("target_file", "") or "")
        return bool(target and open_file(target))

    def open_record_folder(self, record: dict) -> bool:
        target = str(record.get("target_file", "") or "")
        return bool(target and open_folder(target))

    def get_selected_records(self) -> list[dict]:
        seen: set[int] = set()
        rows: list[dict] = []
        for item in self.selectedItems():
            row = item.row()
            if row in seen:
                continue
            seen.add(row)
            record_item = self.item(row, 0)
            if not record_item:
                continue
            record = record_item.data(Qt.ItemDataRole.UserRole)
            if record:
                rows.append(record)
        return rows

    def get_filtered_records(self) -> list[dict]:
        rows: list[dict] = []
        for row in range(self.rowCount()):
            record_item = self.item(row, 0)
            if not record_item:
                continue
            record = record_item.data(Qt.ItemDataRole.UserRole)
            if record:
                rows.append(record)
        return rows

    def get_effective_records(self) -> list[dict]:
        selected = self.get_selected_records()
        return selected if selected else self.get_filtered_records()
