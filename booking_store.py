"""Persistente Buchungsbasis fuer Verlauf, Dateiaktionen und Rerun."""
from __future__ import annotations

import json
from dataclasses import asdict
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from models import CollectionRecord, PartnerInfo


STORE_HEADERS = [
    "entry_id",
    "subject",
    "received",
    "booking_date",
    "amount",
    "purpose",
    "scope",
    "template_kind",
    "booking_type",
    "aobj",
    "sachkonto",
    "partner",
    "source_text",
    "needs_review",
    "match_reason",
    "target_file",
    "status",
    "updated_at",
]


def store_path(cfg: dict) -> Path:
    return Path(cfg["state"].get("booking_store_file") or (Path(__file__).parent / "data" / "state" / "bookings.json"))


def load_bookings(cfg: dict) -> list[dict[str, Any]]:
    path = store_path(cfg)
    if not path.exists():
        migrated = _migrate_from_overview(cfg)
        if migrated:
            save_bookings(cfg, migrated)
            return migrated
        return []
    with path.open(encoding="utf-8") as f:
        raw = json.load(f)
    if not isinstance(raw, list):
        return []
    rows: list[dict[str, Any]] = []
    for item in raw:
        if not isinstance(item, dict):
            continue
        rows.append(_normalize_row(item))
    return rows


def save_bookings(cfg: dict, rows: Iterable[dict[str, Any]]) -> None:
    path = store_path(cfg)
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = [_serialize_row(_normalize_row(dict(row))) for row in rows]
    with path.open("w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)


def upsert_booking(cfg: dict, record: CollectionRecord, target_file: str, status: str = "ok") -> dict[str, Any]:
    rows = load_bookings(cfg)
    row = record.to_overview_row()
    row["partner"] = asdict(record.partner) if record.partner else {}
    row["target_file"] = target_file
    row["status"] = status
    row["updated_at"] = datetime.now().isoformat()
    row = _normalize_row(row)
    replaced = False
    for idx, existing in enumerate(rows):
        if existing.get("entry_id") == record.entry_id:
            rows[idx] = row
            replaced = True
            break
    if not replaced:
        rows.append(row)
    save_bookings(cfg, rows)
    return row


def remove_bookings(cfg: dict, entry_ids: set[str]) -> list[dict[str, Any]]:
    rows = load_bookings(cfg)
    removed = [row for row in rows if str(row.get("entry_id") or "") in entry_ids]
    kept = [row for row in rows if str(row.get("entry_id") or "") not in entry_ids]
    save_bookings(cfg, kept)
    return removed


def get_booking_rows(cfg: dict, *, include_statuses: set[str] | None = None) -> list[dict[str, Any]]:
    rows = load_bookings(cfg)
    if include_statuses is None:
        return rows
    return [row for row in rows if str(row.get("status") or "") in include_statuses]


def row_to_record(row: dict[str, Any]) -> CollectionRecord:
    partner_raw = row.get("partner") or {}
    partner = None
    if isinstance(partner_raw, dict) and any(str(v or "").strip() for v in partner_raw.values()):
        partner = PartnerInfo(**{k: str(v or "") for k, v in partner_raw.items()})
    received = row.get("received")
    if isinstance(received, str) and received:
        received = datetime.fromisoformat(received)
    booking_date = row.get("booking_date")
    if isinstance(booking_date, str):
        booking_date = date.fromisoformat(booking_date[:10])
    return CollectionRecord(
        entry_id=str(row.get("entry_id") or ""),
        subject=str(row.get("subject") or ""),
        received=received if isinstance(received, datetime) else None,
        booking_date=booking_date,
        amount=float(row.get("amount") or 0),
        purpose=str(row.get("purpose") or ""),
        scope=str(row.get("scope") or ""),
        template_kind=str(row.get("template_kind") or ""),
        booking_type=str(row.get("booking_type") or ""),
        aobj=str(row.get("aobj") or ""),
        sachkonto=str(row.get("sachkonto") or ""),
        partner=partner,
        source_text=str(row.get("source_text") or ""),
        needs_review=bool(row.get("needs_review", False)),
        match_reason=str(row.get("match_reason") or ""),
    )


def _normalize_row(row: dict[str, Any]) -> dict[str, Any]:
    normalized = {key: row.get(key, "") for key in STORE_HEADERS}
    for key in ("entry_id", "subject", "purpose", "scope", "template_kind", "booking_type", "aobj", "sachkonto", "source_text", "match_reason", "target_file", "status", "updated_at"):
        normalized[key] = str(normalized.get(key) or "")
    received = normalized.get("received")
    if isinstance(received, datetime):
        normalized["received"] = received.isoformat()
    elif not isinstance(received, str):
        normalized["received"] = ""
    booking_date = normalized.get("booking_date")
    if isinstance(booking_date, datetime):
        normalized["booking_date"] = booking_date.date()
    elif isinstance(booking_date, str) and booking_date:
        normalized["booking_date"] = date.fromisoformat(booking_date[:10])
    elif not isinstance(booking_date, date):
        normalized["booking_date"] = date.today()
    try:
        normalized["amount"] = float(normalized.get("amount") or 0)
    except Exception:
        normalized["amount"] = 0.0
    normalized["needs_review"] = bool(normalized.get("needs_review", False))
    partner = normalized.get("partner") or {}
    normalized["partner"] = partner if isinstance(partner, dict) else {}
    return normalized


def _serialize_row(row: dict[str, Any]) -> dict[str, Any]:
    payload = dict(row)
    booking_date = payload.get("booking_date")
    if isinstance(booking_date, date):
        payload["booking_date"] = booking_date.isoformat()
    return payload


def _migrate_from_overview(cfg: dict) -> list[dict[str, Any]]:
    try:
        import openpyxl

        overview_path = Path(cfg["output"]["overview_file"])
        if not overview_path.exists():
            return []
        wb = openpyxl.load_workbook(overview_path, read_only=True, data_only=True)
        if "Uebersicht" not in wb.sheetnames:
            wb.close()
            return []
        ws = wb["Uebersicht"]
        headers = [str(ws.cell(1, c).value or "") for c in range(1, ws.max_column + 1)]
        rows: list[dict[str, Any]] = []
        for values in ws.iter_rows(min_row=2, values_only=True):
            if not any(values):
                continue
            row = dict(zip(headers, values))
            row["partner"] = {
                "partner_nr": str(row.get("partner_nr") or ""),
                "name_institution": str(row.get("partner_name") or ""),
                "anschrift": "",
                "bankname": str(row.get("partner_bankname") or ""),
                "iban": str(row.get("partner_iban") or ""),
                "bic": str(row.get("partner_bic") or ""),
            }
            row["status"] = "ok"
            rows.append(_normalize_row(row))
        wb.close()
        return rows
    except Exception:
        return []
