"""Excel-Schreiber fuer Monatsdateien und Uebersicht."""
from __future__ import annotations

import logging
import shutil
from pathlib import Path

import openpyxl

from overview import upsert_overview_record

log = logging.getLogger(__name__)

OWN_SHEET_NAME = "eigene Gemeinde"
FORWARD_SHEET_NAME = "zur Weiterleitung"
DATA_START_ROW = 17
OWN_MAX_ROW = 30
FORWARD_MAX_ROW = 21


def write_collection(record, cfg: dict) -> Path:
    """Schreibt eine klassifizierte Kollekte in die passende Monatsdatei."""
    if record.template_kind == "zur_weiterleitung":
        return _write_forwarding(record, cfg)
    return _write_own(record, cfg)


def rebuild_outputs(records: list, cfg: dict) -> None:
    root = Path(cfg["output"]["root_dir"])
    for kind in ("eigene_gemeinde", "zur_weiterleitung"):
        target_dir = root / kind
        target_dir.mkdir(parents=True, exist_ok=True)
        for existing in target_dir.glob("Kollekten_*.xlsx"):
            existing.unlink()
    overview_path = Path(cfg["output"]["overview_file"])
    if overview_path.exists():
        overview_path.unlink()
    for record in sorted(records, key=lambda item: (item.booking_date, item.entry_id)):
        write_collection(record, cfg)


def _write_own(record, cfg: dict) -> Path:
    template = Path(cfg["templates"]["eigene_gemeinde"])
    monthly_path = _get_monthly_path(record.booking_date, cfg, "eigene_gemeinde", template)
    wb = openpyxl.load_workbook(monthly_path)
    if OWN_SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Sheet '{OWN_SHEET_NAME}' nicht gefunden in {monthly_path}.")
    ws = wb[OWN_SHEET_NAME]
    row = _find_next_empty_row(ws, OWN_MAX_ROW)
    _write_header_info(ws, cfg)
    ws.cell(row=row, column=1, value=record.amount).number_format = "#,##0.00 €"
    if record.aobj:
        ws.cell(row=row, column=2, value=record.aobj)
    if record.sachkonto:
        ws.cell(row=row, column=4, value=record.sachkonto)
    ws.cell(row=row, column=6, value=record.booking_date).number_format = "DD.MM.YYYY"
    ws.cell(row=row, column=8, value=record.purpose)
    wb.save(monthly_path)
    log.info("Eingetragen in %s Zeile %d", monthly_path.name, row)
    upsert_overview_record(cfg["output"]["overview_file"], record, str(monthly_path))
    return monthly_path


def _write_forwarding(record, cfg: dict) -> Path:
    template = Path(cfg["templates"]["zur_weiterleitung"])
    monthly_path = _get_monthly_path(record.booking_date, cfg, "zur_weiterleitung", template)
    wb = openpyxl.load_workbook(monthly_path)
    if FORWARD_SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Sheet '{FORWARD_SHEET_NAME}' nicht gefunden in {monthly_path}.")
    ws = wb[FORWARD_SHEET_NAME]
    row = _find_next_empty_row(ws, FORWARD_MAX_ROW)
    _write_header_info(ws, cfg)
    ws.cell(row=row, column=1, value=record.amount).number_format = "#,##0.00 €"
    ws.cell(row=row, column=6, value=record.booking_date).number_format = "DD.MM.YYYY"
    ws.cell(row=row, column=8, value=record.purpose)
    ws.cell(row=row, column=12, value=record.booking_type or "Kollekten")
    partner = record.partner
    if partner and not partner.is_empty():
        ws["C25"] = partner.partner_nr
        ws["C26"] = partner.name_institution
        ws["C27"] = partner.anschrift
        ws["C28"] = partner.bankname
        ws["C29"] = partner.iban
        ws["C31"] = partner.bic
    wb.save(monthly_path)
    log.info("Eingetragen in %s Zeile %d", monthly_path.name, row)
    upsert_overview_record(cfg["output"]["overview_file"], record, str(monthly_path))
    return monthly_path


def _get_monthly_path(datum, cfg: dict, kind: str, template: Path) -> Path:
    out = Path(cfg["output"]["root_dir"]) / kind
    out.mkdir(parents=True, exist_ok=True)
    filename = f"Kollekten_{datum.year}_{datum.month:02d}.xlsx"
    dest = out / filename
    if not dest.exists():
        shutil.copy2(template, dest)
        _write_template_header(dest, cfg)
        log.info("Neue Monatsdatei erstellt: %s", dest)
    return dest


def _write_template_header(path: Path, cfg: dict) -> None:
    wb = openpyxl.load_workbook(path)
    for sheet_name in (OWN_SHEET_NAME, FORWARD_SHEET_NAME):
        if sheet_name in wb.sheetnames:
            _write_header_info(wb[sheet_name], cfg)
    wb.save(path)


def _write_header_info(ws, cfg: dict) -> None:
    organization = cfg.get("organization", {})
    if organization.get("rechtsträger_nr"):
        ws["L2"] = organization["rechtsträger_nr"]
    if organization.get("bank_name"):
        ws["B11"] = organization["bank_name"]
    if organization.get("bank_iban"):
        ws["B12"] = organization["bank_iban"]


def _find_next_empty_row(ws, max_row: int) -> int:
    row = DATA_START_ROW
    while row <= max_row:
        if ws.cell(row=row, column=1).value is None and ws.cell(row=row, column=6).value is None:
            return row
        row += 1
    raise RuntimeError(f"Keine freie Zeile mehr im Blatt '{ws.title}' gefunden.")
