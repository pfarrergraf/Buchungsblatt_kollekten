"""Fortlaufende Gesamtübersicht als separate XLSX-Arbeitsmappe."""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl import Workbook

from models import CollectionRecord


OVERVIEW_HEADERS = [
    "entry_id",
    "subject",
    "received",
    "booking_date",
    "amount",
    "purpose",
    "scope",
    "template_kind",
    "booking_type",
    "aobj",
    "sachkonto",
    "partner_nr",
    "partner_name",
    "partner_iban",
    "partner_bic",
    "partner_bankname",
    "needs_review",
    "match_reason",
    "source_text",
    "target_file",
    "updated_at",
]


def ensure_overview_workbook(path: str | Path) -> Path:
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    if p.exists():
        return p
    wb = Workbook()
    ws = wb.active
    ws.title = "Uebersicht"
    ws.append(OVERVIEW_HEADERS)
    wb.save(p)
    return p


def upsert_overview_record(path: str | Path, record: CollectionRecord, target_file: str) -> None:
    p = ensure_overview_workbook(path)
    wb = openpyxl.load_workbook(p)
    if "Uebersicht" not in wb.sheetnames:
        ws = wb.create_sheet("Uebersicht")
        ws.append(OVERVIEW_HEADERS)
    else:
        ws = wb["Uebersicht"]
    header_map = {str(ws.cell(1, c).value): c for c in range(1, ws.max_column + 1)}
    entry_col = header_map["entry_id"]
    row = _find_row_by_entry_id(ws, entry_col, record.entry_id)
    values = record.to_overview_row()
    values["target_file"] = target_file
    values["updated_at"] = datetime.now()
    # Strip timezone info — openpyxl does not support tz-aware datetimes
    for key, val in values.items():
        if isinstance(val, datetime) and val.tzinfo is not None:
            values[key] = val.replace(tzinfo=None)
    row_values = [values.get(header, "") for header in OVERVIEW_HEADERS]
    if row is None:
        ws.append(row_values)
    else:
        for idx, value in enumerate(row_values, start=1):
            ws.cell(row=row, column=idx, value=value)
    wb.save(p)


def _find_row_by_entry_id(ws, entry_col: int, entry_id: str) -> int | None:
    for row in range(2, ws.max_row + 1):
        if str(ws.cell(row=row, column=entry_col).value or "") == entry_id:
            return row
    return None
